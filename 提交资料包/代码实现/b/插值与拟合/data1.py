import pandas as pd
import numpy as np
from scipy.interpolate import griddata
from openpyxl import Workbook
#输入附件原始海域散点数据，输出原始坐标与法向量
data = pd.read_excel('附件.xlsx')
x1 = data.iloc[0, 2:].values
y1 = data.iloc[1:, 1].values
x1 = x1 * 1852
y1 = y1 * 1852
x = np.zeros(len(x1) * len(y1))
y = np.zeros(len(x1) * len(y1))
z = np.zeros(len(x1) * len(y1))
z1 = np.zeros((len(x1), len(y1)))
i = 0
j = 0
for i in range(0, len(y1)):
    for j in range(0, len(x1)):
        z1[j, i] = data.iloc[i + 1, j + 2] * -1
for i in range(0, len(y1)):
    for j in range(0, len(x1)):
        x[i * len(x1) + j] = x1[j]
        y[i * len(x1) + j] = y1[i]
        z[i * len(x1) + j] = z1[j, i]
z = np.array(z)
x = np.array(x)
y = np.array(y)

xi = np.linspace(x.min(), x.max(), 201 * 4)
yi = np.linspace(y.min(), y.max(), 251 * 4)
xi, yi = np.meshgrid(xi, yi)
zi = griddata((x, y), z, (xi, yi), method='cubic')
#print(zi.shape)
# 在指定坐标处计算法向量
# x0 = 50
# y0 = 50
# 计算指定坐标处的法向量
dxi, dyi = np.gradient(zi)
print(xi.shape)
print(yi.shape)
# xt = [1, 0, dxi[y0, x0]]
# yt = [0, 1, dyi[y0, x0]]
# normal = np.cross(xt, yt)
# print(normal)
# xt = [1, 0, 0]
# yt = [0, 1, 0]
workbook = Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = 'x'
sheet.cell(row=1, column=2).value = 'y'
sheet.cell(row=1, column=3).value = 'z'
sheet.cell(row=1, column=4).value = 'xn'
sheet.cell(row=1, column=5).value = 'yn'
sheet.cell(row=1, column=6).value = 'zn'
for i in range(0, len(y1)):
    for j in range(0, len(x1)):
        index = i * len(x1) + j + 2
        sheet.cell(row=index, column=1).value = x1[j]
        sheet.cell(row=index, column=2).value = y1[i]
        sheet.cell(row=index, column=3).value = z1[j, i]
        x0 = j * 4  # 需要计算法向量的 x 坐标
        y0 = i * 4 # 需要计算法向量的 y 坐标
        # 计算指定坐标处的法向量
        xt = [1, 0, dxi[y0, x0]]
        yt = [0, 1, dyi[y0, x0]]
        normal = np.cross(xt, yt)
        normal = normal / np.linalg.norm(normal)
        sheet.cell(row=index, column=4).value = normal[0]
        sheet.cell(row=index, column=5).value = normal[1]
        sheet.cell(row=index, column=6).value = normal[2]
workbook.save('data1.xlsx')
workbook.close()
