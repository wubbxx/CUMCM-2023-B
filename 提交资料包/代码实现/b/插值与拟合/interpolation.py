import pandas as pd
import numpy as np
from scipy.interpolate import griddata
from openpyxl import Workbook

data = pd.read_excel('附件.xlsx')
x1 = data.iloc[0, 2:].values
y1 = data.iloc[1:, 1].values
x1 = x1 * 1852
y1 = y1 * 1852
x = np.zeros(len(x1) * len(y1))
y = np.zeros(len(x1) * len(y1))
z = np.zeros(len(x1) * len(y1))
z1 = np.zeros((len(x1), len(y1)))
i = 0
j = 0
for i in range(0, len(y1)):
    for j in range(0, len(x1)):
        z1[j, i] = data.iloc[i + 1, j + 2] * -1
for i in range(0, len(y1)):
    for j in range(0, len(x1)):
        x[i * len(x1) + j] = x1[j]
        y[i * len(x1) + j] = y1[i]
        z[i * len(x1) + j] = z1[j, i]
z = np.array(z)
x = np.array(x)
y = np.array(y)

xi = np.linspace(x.min(), x.max(), 201 * 4)
yi = np.linspace(y.min(), y.max(), 251 * 4)
xi, yi = np.meshgrid(xi, yi)
zi = griddata((x, y), z, (xi, yi), method='cubic')
print(zi.shape)
workbook = Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = 'x'
sheet.cell(row=1, column=2).value = 'y'
sheet.cell(row=1, column=3).value = 'z'
for i in range(0, 4 * len(y1)):
    for j in range(0, 4 * len(x1)):
        index = i * 4 * len(x1) + j + 2
        sheet.cell(row=index, column=1).value = xi[i][j]
        sheet.cell(row=index, column=2).value = yi[i][j]
        sheet.cell(row=index, column=3).value = zi[i, j]
workbook.save('interpolation.xlsx')
workbook.close()