import numpy as np
from scipy.optimize import least_squares


# 误差函数可以定义点到平面的距离，适用数据中不存在较大偏差和离群值
def residual(params, x, y, z):
    a, b, c = params
    x = np.array(x)
    y = np.array(y)
    z = np.array(z)
    dist = abs((a*x + b*y + c - z) / math.sqrt(a*a + b*b + 1))
    return dist

#残差平方和 适用较大偏差
def ssr(params, x, y, z):
    a, b, c = params
    x = np.array(x)
    y = np.array(y)
    z = np.array(z)
    return a*x + b*y + c - z

def fitting(x, y, z, model):
# 初始化参数估计值
    params0 = np.array([0.047201027, -0.019578743, 0.998693515])

# 使用最小二乘法进行拟合
    if model == 1:
        result = least_squares(residual, params0, args=(x, y, z))
    else:
        result = least_squares(ssr, params0, args=(x, y, z))

# 获取拟合结果
    a, b, c = result.x
    print("拟合平面方程：z = {:.6f}*x + {:.6f}*y + {:.6f}".format(a, b, c))
    return [a, b, c]


# 适用于四边形
def quadrilateral(x, y, z):
    result = fitting(x ,y, z, 1)
    normal = (result[0], result[1], -1)
    vector_hor = (result[0], result[1])
    vector_hor = vector_hor / np.linalg.norm(vector_hor)

import pandas as pd
from openpyxl import Workbook
import math 
from sklearn.metrics import r2_score
# 读取 Excel 文件
#读取 Excel 文件
df = pd.read_excel('output.xlsx')

# 按照多列关键词进行排序
sorted_df = df.sort_values(by=['Label', 'Dim2', 'Dim1'])

# 重置索引
sorted_df.reset_index(drop=True, inplace=True)

# 将排序后的结果写回到原始 Excel 文件中
sorted_df.to_excel('output.xlsx', index=False)

df = pd.read_excel('output.xlsx')
labels = int(input("输入labels"))
i = 0
my_dict = {}
boder = {}
equations = {}
for i in range(0, labels):
    my_dict[str(i)] = {'x': [], 'y': [], 'z':[]}
    boder[str(i)] = {'l': [], 'r': []}
    equations[str(i)] = {'a': 0, 'b': 0, 'c': 0, 'd': 0}
x = []
y = []
z = []
x1 = df.iloc[0:, 0]
y1 = df.iloc[0:, 1]
z1 = df.iloc[0:, 2]
label1 = df.iloc[0:, 6]
i = 0
for i in range(0, len(x1)):
    my_dict[str(int(label1[i]))]['x'].append(x1[i])
    my_dict[str(int(label1[i]))]['y'].append(y1[i])
    my_dict[str(int(label1[i]))]['z'].append(z1[i])

workbook = Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = 'a'
sheet.cell(row=1, column=2).value = 'b'
sheet.cell(row=1, column=3).value = 'c'
sheet.cell(row=1, column=4).value = 'd'
sheet.cell(row=1, column=5).value = 'normal.x'
sheet.cell(row=1, column=6).value = 'normal.y'
sheet.cell(row=1, column=7).value = 'normal.z'
sheet.cell(row=1, column=8).value = 'alpha'
sheet.cell(row=1, column=9).value = 'label'
sheet.cell(row=1, column=10).value = 'R-squared'
j = 2
for i in range(0, labels):
    x = my_dict[str(i)]['x']
    y = my_dict[str(i)]['y']
    z = my_dict[str(i)]['z']
    result = fitting(x ,y, z, 2)
    equation_str = "{:.6f}*x + {:.6f}*y - z + {:.6f} = 0".format(result[0], result[1], result[2])
    normal = (-1 * result[0], -1 * result[1], 1)
    normal = normal / np.linalg.norm(normal)
    vector_hor = (-1 * result[0], -1 * result[1])
    vector_hor = vector_hor / np.linalg.norm(vector_hor)
    alpha = math.atan(math.sqrt(normal[0] * normal[0] + normal[1] * normal[1]) / normal[2])
    alpha = math.degrees(alpha)
    equations[str(i)]['a'] = result[0]
    equations[str(i)]['b'] = result[1]
    equations[str(i)]['c'] = -1
    equations[str(i)]['d'] = result[2]
    sheet.cell(row = j, column = 1).value = result[0]
    sheet.cell(row = j, column = 2).value = result[1]
    sheet.cell(row = j, column = 3).value = -1
    sheet.cell(row = j, column = 4).value = result[2]
    sheet.cell(row = j, column = 5).value = normal[0]
    sheet.cell(row = j, column = 6).value = normal[1]
    sheet.cell(row = j, column = 7).value = normal[2]
    sheet.cell(row = j, column = 8).value = alpha
    sheet.cell(row = j, column = 9).value = i
    j = j + 1

for i in range(0, labels):
    x = my_dict[str(i)]['x']
    y = my_dict[str(i)]['y']
    z = my_dict[str(i)]['z']
    z_pred = []
    for k in range(0, len(x)):
        temp = equations[str(i)]['a'] * x[k] + equations[str(i)]['b'] * y[k] + equations[str(i)]['d']
        z_pred.append(temp)
    r_squared = r2_score(z, z_pred)
    sheet.cell(row = i + 2 , column = 10).value = r_squared
workbook.save('plane_output.xlsx')
workbook.close()

