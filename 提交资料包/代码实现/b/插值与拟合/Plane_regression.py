import numpy as np
from scipy.optimize import least_squares


# 误差函数可以定义点到平面的距离，适用数据中不存在较大偏差和离群值
def residual(params, x, y, z):
    a, b, c = params
    x = np.array(x)
    y = np.array(y)
    z = np.array(z)
    dist = abs((a*x + b*y + c - z) / math.sqrt(a*a + b*b + 1))
    return dist

#残差平方和 适用较大偏差
def ssr(params, x, y, z):
    a, b, c = params
    x = np.array(x)
    y = np.array(y)
    z = np.array(z)
    return a*x + b*y + c - z

def fitting(x, y, z, model):
# 初始化参数估计值
    params0 = np.array([0.047201027, -0.019578743, 0.998693515])

# 使用最小二乘法进行拟合
    if model == 1:
        result = least_squares(residual, params0, args=(x, y, z))
    else:
        result = least_squares(ssr, params0, args=(x, y, z))

# 获取拟合结果
    a, b, c = result.x
    print("拟合平面方程：z = {:.6f}*x + {:.6f}*y + {:.6f}".format(a, b, c))
    return [a, b, c]


# 适用于四边形
def quadrilateral(x, y, z):
    result = fitting(x ,y, z, 1)
    normal = (result[0], result[1], -1)
    vector_hor = (result[0], result[1])
    vector_hor = vector_hor / np.linalg.norm(vector_hor)

import pandas as pd
from openpyxl import Workbook
import math 

# 读取 Excel 文件
#读取 Excel 文件
df = pd.read_excel('output.xlsx')

# 按照多列关键词进行排序
sorted_df = df.sort_values(by=['Label', 'Dim2', 'Dim1'])

# 重置索引
sorted_df.reset_index(drop=True, inplace=True)

# 将排序后的结果写回到原始 Excel 文件中
sorted_df.to_excel('output.xlsx', index=False)

df = pd.read_excel('output.xlsx')
labels = 4
my_dict = {}
my_dict['0'] = {'x': [], 'y': [], 'z':[]}
my_dict['1'] = {'x': [], 'y': [], 'z':[]}
my_dict['2'] = {'x': [], 'y': [], 'z':[]}
my_dict['3'] = {'x': [], 'y': [], 'z':[]}
boder = {'0': {'l': [], 'r': []}, '1': {'l': [], 'r': []}, '2': {'l': [], 'r': []}, '3': {'l': [], 'r': []}}
x = []
y = []
z = []
x1 = df.iloc[0:, 0]
y1 = df.iloc[0:, 1]
z1 = df.iloc[0:, 2]
label1 = df.iloc[0:, 6]
i = 0
for i in range(0, len(x1)):
    my_dict[str(int(label1[i]))]['x'].append(x1[i])
    my_dict[str(int(label1[i]))]['y'].append(y1[i])
    my_dict[str(int(label1[i]))]['z'].append(z1[i])

workbook = Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = 'a'
sheet.cell(row=1, column=2).value = 'b'
sheet.cell(row=1, column=3).value = 'c'
sheet.cell(row=1, column=4).value = 'd'
sheet.cell(row=1, column=5).value = 'normal.x'
sheet.cell(row=1, column=6).value = 'normal.y'
sheet.cell(row=1, column=7).value = 'normal.z'
sheet.cell(row=1, column=8).value = 'alpha'
sheet.cell(row=1, column=9).value = 'label'
j = 2
for i in range(0, labels):
    x = my_dict[str(i)]['x']
    y = my_dict[str(i)]['y']
    z = my_dict[str(i)]['z']
    result = fitting(x ,y, z, 2)
    equation_str = "{:.6f}*x + {:.6f}*y - z + {:.6f} = 0".format(result[0], result[1], result[2])
    normal = (-1 * result[0], -1 * result[1], 1)
    normal = normal / np.linalg.norm(normal)
    vector_hor = (-1 * result[0], -1 * result[1])
    vector_hor = vector_hor / np.linalg.norm(vector_hor)
    alpha = math.atan(math.sqrt(normal[0] * normal[0] + normal[1] * normal[1]) / normal[2])
    alpha = math.degrees(alpha)
    sheet.cell(row = j, column = 1).value = result[0]
    sheet.cell(row = j, column = 2).value = result[1]
    sheet.cell(row = j, column = 3).value = -1
    sheet.cell(row = j, column = 4).value = result[2]
    sheet.cell(row = j, column = 5).value = normal[0]
    sheet.cell(row = j, column = 6).value = normal[1]
    sheet.cell(row = j, column = 7).value = normal[2]
    sheet.cell(row = j, column = 8).value = alpha
    sheet.cell(row = j, column = 9).value = i
    j = j + 1

workbook.save('plane_output.xlsx')
workbook.close()

