import pandas as pd
from openpyxl import Workbook

#读取 Excel 文件
df = pd.read_excel('output.xlsx')

# 按照多列关键词进行排序
sorted_df = df.sort_values(by=['Label', 'Dim2', 'Dim1'])

# 重置索引
sorted_df.reset_index(drop=True, inplace=True)

# 将排序后的结果写回到原始 Excel 文件中
sorted_df.to_excel('output.xlsx', index=False)

df = pd.read_excel('output.xlsx')
labels = int(input("输入labels"))
i = 0
my_dict = {}
boder = {}
for i in range(0, labels):
    my_dict[str(i)] =  {'x': [], 'y': [], 'z':[]}
    boder[str(i)] = {'l': [], 'r': []}
x = []
y = []
z = []
x1 = df.iloc[0:, 0]
y1 = df.iloc[0:, 1]
z1 = df.iloc[0:, 2]
label1 = df.iloc[0:, 6]
i = 0
for i in range(0, len(x1)):
    my_dict[str(int(label1[i]))]['x'].append(x1[i])
    my_dict[str(int(label1[i]))]['y'].append(y1[i])
    my_dict[str(int(label1[i]))]['z'].append(z1[i])
workbook = Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = 'x'
sheet.cell(row=1, column=2).value = 'y'
sheet.cell(row=1, column=3).value = 'label'
j = 2
for i in range(0, labels):
    x = my_dict[str(i)]['x']
    y = my_dict[str(i)]['y']
    last = y[0]
    index_l = 0
    index_r = 0
    while index_r < len(y):
        while abs(y[index_l] - y[index_r]) < 1e-3:
            index_r = index_r + 1
            if index_r == len(y):
                break
        t_l = (x[index_l], y[index_l])
        t_r = (x[index_r - 1], y[index_r - 1])
        boder[str(i)]['l'].append(t_l)
        boder[str(i)]['r'].append(t_r)
        index_l = index_r
        index_r = index_r + 1

for i in range(0, labels):
    b_l = boder[str(i)]['l']
    b_r = boder[str(i)]['r']
    for k in range(0, len(b_l)):
        sheet.cell(row = j, column = 1).value = b_l[k][0]
        sheet.cell(row = j, column = 2).value = b_l[k][1]
        sheet.cell(row = j, column = 3).value = str(i) + '-l'
        j = j + 1
    for k in range(0, len(b_r)):
        sheet.cell(row = j, column = 1).value = b_r[k][0]
        sheet.cell(row = j, column = 2).value = b_r[k][1]
        sheet.cell(row = j, column = 3).value = str(i) + '-r'
        j = j + 1
workbook.save('border_v.xlsx')
workbook.close()
        
    