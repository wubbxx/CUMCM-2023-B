import numpy as np
from scipy import stats
import pandas as pd
from openpyxl import Workbook
import openpyxl

df = pd.read_excel('linear_input.xlsx')
# sum = int(input("请输入sum"))
i = 0
my_dict = {}
# for i in range(0, sum):
#     my_dict[str(i)] = {'x': [], 'y': []}
x = df.iloc[0:, 0]
y = df.iloc[0:, 1]
label = df.iloc[0:, 2]
for i in range(0, len(x)):
    if str(label[i]) not in my_dict:
        my_dict[str(label[i])] = {'x': [], 'y': []}
    my_dict[str(label[i])]['x'].append(x[i])
    my_dict[str(label[i])]['y'].append(y[i])

workbook = Workbook()
sheet = workbook.active
sheet.cell(row=1, column=1).value = 'a'
sheet.cell(row=1, column=2).value = 'b'
sheet.cell(row=1, column=3).value = 'c'
sheet.cell(row=1, column=4).value = 'label'
j = 2
for key, value in my_dict.items():
    x = value['x']
    y = value['y']   
    slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
    sheet.cell(row=j, column=1).value = slope
    sheet.cell(row=j, column=2).value = -1
    sheet.cell(row=j, column=3).value = intercept
    sheet.cell(row=j, column=4).value = key
    j = j + 1

# 保存修改后的 Excel 文件
workbook.save('linear_output.xlsx')
workbook.close()